from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rest_framework_simplejwt.views import TokenObtainPairView
from .models import Usuario, Planificacion, Asignatura, ModuloAsignatura, CargaDocente, OtraActividad
from .serializers import (
    UsuarioSerializer, PlanificacionSerializer, AsignaturaSerializer,
    ModuloAsignaturaSerializer, CargaDocenteSerializer, OtraActividadSerializer,
    CustomTokenObtainPairSerializer
)
from .calculos import calcular_resumen_docente, calcular_resumen_planificacion
from .permissions import EsCoordinador


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

    def get_queryset(self):
        if self.request.user.rol == "coordinador":
            return Usuario.objects.all()
        return Usuario.objects.filter(id=self.request.user.id)
    
    @action(detail=False, methods=["get"])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)


class PlanificacionViewSet(viewsets.ModelViewSet):
    queryset = Planificacion.objects.all()
    serializer_class = PlanificacionSerializer

    # ← Resumen PERSONAL (para el Dashboard, funciona para cualquier rol)
    @action(detail=True, methods=["get"], url_path="mi-resumen")
    def mi_resumen(self, request, pk=None):
        planificacion = self.get_object()
        data = calcular_resumen_docente(request.user, planificacion)
        return Response(data)

    # ← Resumen CONSOLIDADO (solo coordinadores, para /resumen)
    @action(detail=True, methods=["get"], url_path="resumen")
    def resumen(self, request, pk=None):
        planificacion = self.get_object()
        if request.user.rol != "coordinador":
            return Response(
                {"detail": "No tienes permiso para ver el resumen consolidado."},
                status=403
            )
        data = calcular_resumen_planificacion(planificacion)
        return Response(data)

    @action(detail=True, methods=["get"])
    def exportar_excel(self, request, pk=None):
        planificacion = self.get_object()
        resumen = calcular_resumen_planificacion(planificacion)
        wb = _generar_excel(planificacion, resumen)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="PLANAC_{planificacion.anio}_{planificacion.semestre}.xlsx"'
        )
        wb.save(response)
        return response


class AsignaturaViewSet(viewsets.ModelViewSet):
    queryset = Asignatura.objects.prefetch_related("modulos").all()
    serializer_class = AsignaturaSerializer


class ModuloAsignaturaViewSet(viewsets.ModelViewSet):
    queryset = ModuloAsignatura.objects.select_related("asignatura").all()
    serializer_class = ModuloAsignaturaSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        asignatura_id = self.request.query_params.get("asignatura")
        if asignatura_id:
            qs = qs.filter(asignatura_id=asignatura_id)
        return qs


class CargaDocenteViewSet(viewsets.ModelViewSet):
    queryset = CargaDocente.objects.select_related("modulo__asignatura", "docente").all()
    serializer_class = CargaDocenteSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        planificacion_id = self.request.query_params.get("planificacion")
        docente_id = self.request.query_params.get("docente")

        # Docente solo ve las suyas; coordinador puede ver cualquiera
        if self.request.user.rol != "coordinador":
            qs = qs.filter(docente=self.request.user)
        elif docente_id:
            qs = qs.filter(docente_id=docente_id)

        if planificacion_id:
            qs = qs.filter(planificacion_id=planificacion_id)
        return qs

    def perform_create(self, serializer):
        # Coordinador puede crear para cualquier docente; docente solo para sí mismo
        if self.request.user.rol == "coordinador" and "docente" in self.request.data:
            serializer.save()
        else:
            serializer.save(docente=self.request.user)

    @action(detail=False, methods=["post"])
    def bulk_update(self, request):
        cargas_data = request.data.get("cargas", [])
        resultados = []
        for item in cargas_data:
            # Coordinador puede especificar cualquier docente_id
            if request.user.rol == "coordinador" and "docente" in item:
                docente_id = item["docente"]
            else:
                docente_id = request.user.id

            obj, _ = CargaDocente.objects.update_or_create(
                planificacion_id=item["planificacion"],
                docente_id=docente_id,
                modulo_id=item["modulo"],
                defaults={
                    "cantidad": item["cantidad"],
                    "observaciones": item.get("observaciones", ""),
                },
            )
            resultados.append(CargaDocenteSerializer(obj).data)
        return Response(resultados, status=status.HTTP_200_OK)


class OtraActividadViewSet(viewsets.ModelViewSet):
    queryset = OtraActividad.objects.all()
    serializer_class = OtraActividadSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        planificacion_id = self.request.query_params.get("planificacion")
        docente_id = self.request.query_params.get("docente")

        if self.request.user.rol != "coordinador":
            qs = qs.filter(docente=self.request.user)
        elif docente_id:
            qs = qs.filter(docente_id=docente_id)

        if planificacion_id:
            qs = qs.filter(planificacion_id=planificacion_id)
        return qs

    def perform_create(self, serializer):
        if self.request.user.rol == "coordinador" and "docente" in self.request.data:
            serializer.save()
        else:
            serializer.save(docente=self.request.user, planificacion_id=self.request.data.get("planificacion"))


def _generar_excel(planificacion, resumen):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen PLANAC"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    exceso_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    deficit_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    headers = ["Docente", "Horas Contrato", "Total Pregrado", "Otras Actividades", "Total General", "Diferencia", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, d in enumerate(resumen, 2):
        ws.cell(row=row, column=1, value=d["docente"])
        ws.cell(row=row, column=2, value=d["horas_contrato"])
        ws.cell(row=row, column=3, value=d["total_pregrado"])
        ws.cell(row=row, column=4, value=d["total_otras_actividades"])
        ws.cell(row=row, column=5, value=d["total_general"])
        diff_cell = ws.cell(row=row, column=6, value=d["diferencia"])
        estado_cell = ws.cell(row=row, column=7, value=d["estado"])
        fill = exceso_fill if d["estado"] == "exceso" else deficit_fill if d["estado"] == "deficit" else ok_fill
        diff_cell.fill = fill
        estado_cell.fill = fill

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20

    return wb
